"""
make_parameter_dictionary.py — generates output/parameter_dictionary.xlsx:
the COMPLETE dictionary of every parameter the engine uses — which source
Excel sheet and column each figure comes from, every derived field with its
formula, and every model constant with its role and rationale.

Run:  python make_parameter_dictionary.py
openpyxl is used here only (build tool, not the runtime core).
"""

import os
import sys
from datetime import datetime, timezone

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "output", "parameter_dictionary.xlsx")

HDR = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill("solid", fgColor="123A63")
SEC = Font(bold=True, size=11, color="123A63")
WRAP = Alignment(wrap_text=True, vertical="top")


def sheet(wb, title, headers, widths, rows, intro=None):
    ws = wb.create_sheet(title)
    r = 1
    if intro:
        ws.cell(r, 1, intro).font = SEC
        ws.cell(r, 1).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=len(headers))
        ws.row_dimensions[r].height = 34
        r += 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = HDR
        cell.fill = HDR_FILL
    ws.freeze_panes = ws.cell(r + 1, 1)
    for row in rows:
        r += 1
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = WRAP
            cell.font = Font(size=10)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ---- 1 · READ ME -----------------------------------------------------
    ws = wb.create_sheet("READ ME")
    lines = [
        ("Parameter Dictionary — MSME Comparable Valuation Engine", SEC),
        (f"Generated {datetime.now(timezone.utc).isoformat(timespec='seconds')} · "
         f"methodology v2.0.0 · regenerate with: python make_parameter_dictionary.py", None),
        ("", None),
        ("HOW TO READ THIS WORKBOOK", SEC),
        ("1 · Source data:   every column read from the 9 uploaded Excel extracts — "
         "file, column position, header, database field, and what each figure is used for.", None),
        ("2 · Derived fields:   every value COMPUTED from the source columns, with its "
         "exact formula and where it feeds the model.", None),
        ("3 · Model parameters:   every constant in the engine (weights, thresholds, "
         "discounts, anchors) with its value, code location, role and rationale.", None),
        ("4 · Calculation flow:   the end-to-end chain from a company name to the "
         "equity-value range, step by step.", None),
        ("", None),
        ("UNITS: source figures are INR Crore in the extracts; the client emits the D&B "
         "convention (INR Thousand, ×10,000) and normalization converts back to Crore "
         "(÷10,000) — all model math is in INR Crore.", None),
        ("HONESTY: the extracts contain NO market prices, borrowings, cash or current "
         "liabilities. Consequences (all disclosed on every result): book capital "
         "employed ≈ net worth is the EV proxy, re-levelled to sector trading anchors "
         "(sheet 3); net debt is assumed 0 with a warning when unknown.", None),
    ]
    for i, (txt, f) in enumerate(lines, 1):
        ws.cell(i, 1, txt).font = f or Font(size=10)
        ws.cell(i, 1).alignment = WRAP
    ws.column_dimensions["A"].width = 120

    # ---- 2 · Source data ---------------------------------------------------
    src = [
        # file, col, header, db field, used for
        ("Basic Data.xlsx", "B (1)", "Accord Code", "companies.accord",
         "primary key joining every sheet; becomes the DUNS-equivalent identifier"),
        ("Basic Data.xlsx", "C (2)", "Company Name", "companies.name",
         "search / resolution; report labelling"),
        ("Basic Data.xlsx", "D (3)", "CD_Industry", "companies.industry",
         "industry dimension: mapped to a stable pseudo-NAICS per category (similarity 1.0), "
         "a keyword sector GROUP (similarity 0.6, sector anchors), and a D&B major letter "
         "(hard knock-out filter #5)"),
        ("Basic Data.xlsx", "E (4)", "CD_Economic Activity(NIC)", "companies.nic",
         "fallback text for the economic classifier when description is empty"),
        ("Basic Data.xlsx", "F (5)", "CD_Business Description", "companies.description",
         "drives the rule-based economic classifier: operating model (manufacturer/"
         "distributor/retailer/service), value chain, customer type — filters #3, #4 and "
         "similarity #9"),
        ("Basic Data.xlsx", "G (6)", "CD_CIN Number", "companies.cin",
         "listing status: CIN first letter 'L' = listed, 'U' = unlisted → DLOM decision; "
         "identity on the report"),
        ("Basic Data.xlsx", "H (7)", "CD_ISIN No", "companies.isin",
         "stored (future market-price feed join key); not used in calculations yet"),
        ("PL data.xlsx", "A (0)", "Accord Code", "fin.accord", "join key"),
        ("PL data.xlsx", "C (2)", "PL_Year End", "fin.year_end",
         "fiscal-year key; latest + prior 12-month years kept per company"),
        ("PL data.xlsx", "F (5)", "PL_No of Months", "fin.months",
         "period-length gate: only 12-month periods are valuation-grade"),
        ("PL data.xlsx", "BP (67)", "PL_Net Sales", "fin.revenue",
         "REVENUE — EV/Revenue driver; scale similarity #7; DLOM size band; size factor "
         "on sector anchors; scale-mismatch penalty #15; growth calculation"),
        ("PL data.xlsx", "IN (247)", "PL_Operating Profit (Excl OI)", "fin.ebitda",
         "EBITDA — EV/EBITDA driver (headline); EBITDA margin → quality positioning + "
         "margin similarity #8; P&L reconciliation gate"),
        ("PL data.xlsx", "IO (248)", "PL_Other Income", "fin.other_income",
         "P&L reconciliation gate only (EBITDA + OI − Interest ≈ PBDT)"),
        ("PL data.xlsx", "JV (281)", "PL_Interest", "fin.interest",
         "P&L reconciliation gate; reported in company detail"),
        ("PL data.xlsx", "KG (292)", "PL_PBDT", "fin.pbdt",
         "P&L reconciliation gate target value"),
        ("PL data.xlsx", "KH (293)", "PL_Depreciation", "fin.depreciation",
         "EBIT = EBITDA − depreciation → EV/EBIT driver (third triangulation method)"),
        ("PL data.xlsx", "KS (304)", "PL_Profit Before Tax", "fin.pbt",
         "emitted in the D&B envelope; reserved for a future P/E method"),
        ("PL data.xlsx", "LE (316)", "PL_Profit After Tax", "fin.pat",
         "emitted in the D&B envelope; reserved for a future P/E method"),
        ("BS data.xlsx", "B (1)", "Accord Code", "fin.accord (join)", "join key"),
        ("BS data.xlsx", "D (3)", "BS_Year End", "fin.year_end (join)", "join key"),
        ("BS data.xlsx", "E (4)", "Plant & Machinery", "fin.plant_mach",
         "stored; asset-intensity context (not yet a filter)"),
        ("BS data.xlsx", "F (5)", "BS_Net Block", "fin.net_block",
         "tangible fixed assets in the company detail"),
        ("BS data.xlsx", "G (6)", "BS_Inventories", "fin.inventories",
         "inventory in the company detail"),
        ("BS data.xlsx", "H (7)", "BS_Total Assets", "fin.total_assets",
         "total assets in the company detail / data-quality context"),
        ("Net worth.xlsx", "B (1)", "Accord Code", "fin.accord (join)", "join key"),
        ("Net worth.xlsx", "D (3)", "FH_Year End", "fin.year_end (join)", "join key"),
        ("Net worth.xlsx", "E (4)", "FH_Net Worth", "fin.net_worth",
         "NET WORTH — the book EV proxy when no segment capital employed exists "
         "(then re-levelled by sector anchors); valuation-grade criterion (must be > 0)"),
        ("Forex.xlsx", "B (1)", "Accord Code", "fin.accord (join)", "join key"),
        ("Forex.xlsx", "D (3)", "FX_Year_End", "fin.year_end (join)", "join key"),
        ("Forex.xlsx", "E (4)", "FX_Total Inflow In Foreign Currency", "fin.fx_inflow",
         "EXPORTER FLAG: inflow > 0 → is_exporter → export similarity #10"),
        ("Segment.xlsx", "B (1)", "Accord Code", "fin.accord (join)", "join key"),
        ("Segment.xlsx", "D (3)", "FSG_Year End", "fin.year_end (join)", "join key"),
        ("Segment.xlsx", "O (14)", "FSG_Capital employed", "fin.seg_ce (summed per year)",
         "CAPITAL EMPLOYED (preferred book EV proxy when reported; sums across segments)"),
        ("Product.xlsx", "—", "(all columns)", "not loaded",
         "stored with the project; not consumed yet (future product-mix similarity)"),
        ("R&D.xlsx", "—", "(all columns)", "not loaded",
         "stored; not consumed yet (future R&D-intensity signal)"),
        ("Shareholding.xlsx", "—", "(all columns)", "not loaded",
         "stored; not consumed yet (future promoter-holding / free-float signal)"),
    ]
    sheet(wb, "2 · Source data",
          ["Source file", "Excel column", "Header", "Database field",
           "Used for (calculation / filter)"],
          [16, 10, 30, 24, 78], src,
          "Every column the ETL reads from the uploaded extracts. Headers are VERIFIED "
          "before any read — a changed layout aborts the build rather than mis-mapping "
          "figures. Every loaded row stores its source file + Excel row number "
          "(provenance), which is what the report's Data Lineage card shows.")

    # ---- 3 · Derived fields -----------------------------------------------
    der = [
        ("revenue_cr / ebitda_cr / …", "source value ÷ 10,000 (Thousand → Crore)",
         "normalize_company (core/pipeline.py)",
         "all model math is in INR Crore; the client emits D&B-convention Thousand"),
        ("ebit_cr", "ebitda_cr − depreciation_cr (None if either missing)",
         "normalize_company", "EV/EBIT driver — third triangulation method"),
        ("ebitda_margin", "ebitda_cr ÷ revenue_cr",
         "normalize_company", "quality positioning percentile + margin similarity #8 + "
         "data-quality plausibility check (0 < margin < 60%)"),
        ("revenue_growth", "(revenue − prior revenue) ÷ prior revenue",
         "normalize_company", "peer context columns; data-quality LOW check"),
        ("net_debt_cr", "debt_cr − cash_cr; if either UNKNOWN → 0 with NET_DEBT_UNKNOWN "
         "warning", "compute_valuation", "EV → equity bridge"),
        ("capital_employed (book EV proxy)", "segment capital employed when reported, "
         "else NET WORTH", "RealDnBClient._fin_env", "book multiple base, re-levelled "
         "by sector anchors (calibration)"),
        ("is_exporter", "forex inflow > 0", "RealDnBClient._info_env",
         "export similarity #10"),
        ("listed", "CIN[0] == 'L'", "normalize_company", "DLOM 0 for listed; market "
         "cross-check eligibility"),
        ("pseudo-NAICS", "stable 3-digit code per CD_Industry category (sorted order)",
         "RealDnBClient", "industry similarity 1.0 when equal"),
        ("sector group (GRP_*)", "first keyword match over CD_Industry "
         "(21 rule groups: TRADE, RETAIL, FIN, IT, AUTO, PHARMA, CHEM, METAL, TEXTILE, "
         "BUILDMAT, MACH, ELEC, FOOD, POLYPAPER, MINING, ENERGY, CONSTR, LOGIST, MEDIA, "
         "HOTEL, SERVICES)", "realdata/client.py _GROUP_RULES",
         "industry similarity 0.6 when equal; SECTOR ANCHOR lookup for calibration"),
        ("major industry letter", "sector group → letter (D/F/G/I/N/T/C/U)",
         "realdata/client.py _GROUP_MAJOR", "hard knock-out filter #5"),
        ("operating model", "keyword rules over the business description; manufacturing "
         "VERB required (bare noun 'manufacturers' does NOT trigger — collision rule)",
         "build_profile", "hard knock-out filter #3"),
        ("value chain", "'raw material'/'billets'/'supplied to downstream' → raw_material "
         "else finished_goods", "build_profile", "hard knock-out filter #4"),
        ("customer type", "consumer/retail keywords → B2C; oem/industrial/… → B2B; "
         "else mixed", "build_profile", "customer similarity #9"),
        ("valuation-grade flag", "revenue > 0 AND ebitda present AND net worth > 0 AND "
         "12-month period", "etl.py + RealDnBClient", "defines the 14,036-company "
         "usable universe"),
        ("P&L reconciliation", "|EBITDA + Other Income − Interest − PBDT| ≤ "
         "max(2% × |PBDT|, ₹0.5 Cr)", "etl.py", "per-row integrity gate (99.1% pass); "
         "shown in lineage"),
        ("similarity score", "0.40·industry + 0.20·scale + 0.15·margin + 0.15·customer "
         "+ 0.10·export (each dimension raw ∈ [0,1])", "_similarity",
         "peer ranking; match weight; effective peer count"),
        ("match weight", "clamp((score − 0.40)/(0.85 − 0.40), 0.15, 1.0)",
         "_match_weight", "how much each peer's multiple counts (weighted percentile)"),
        ("effective peer count", "Σ match weights of the used peers",
         "compute_valuation", "range-widening trigger; confidence peer coverage"),
        ("quality position (mid_q)", "clamp(percentile rank of target margin among "
         "peers, 0.15, 0.85)", "compute_valuation",
         "the peer-multiple percentile used as the CENTRAL multiple"),
        ("positioned multiple", "weighted percentile of the (calibrated) peer multiples "
         "at mid_q, band at mid_q ± window", "compute_valuation",
         "low/mid/high multiples per method"),
        ("equity per method", "(multiple × driver − net debt) × (1 − DLOM)",
         "compute_valuation", "the answer, as a range"),
        ("confidence score", "0.20·profile + 0.20·coverage + 0.10·EBITDA>0 + "
         "0.10·methods + 0.25·triangulation + 0.15·comp tightness",
         "compute_confidence (run.py)", "HIGH ≥ 0.75 / MEDIUM ≥ 0.50 / LOW"),
    ]
    sheet(wb, "3 · Derived fields",
          ["Derived field", "Formula / rule", "Where computed", "Feeds"],
          [26, 62, 30, 52], der,
          "Every value COMPUTED from the source columns — the exact formula and what "
          "it feeds. Nothing else is derived; if a quantity is not on sheet 2 or here, "
          "the engine does not use it.")

    # ---- 4 · Model parameters ----------------------------------------------
    from core.calibration import SECTOR_ANCHORS, _SIZE_FACTORS, _EBIT_UPLIFT
    par = [
        ("Similarity weight — industry", "0.40", "core/pipeline.py _similarity",
         "same NAICS 3-digit = 1.0 · same sector group = 0.6 · else 0",
         "industry is the strongest determinant of comparability"),
        ("Similarity weight — scale", "0.20", "_similarity",
         "1 / (1 + |log1p(rev_t) − log1p(rev_p)|)",
         "log scale: 10× revenue gap matters the same everywhere"),
        ("Similarity weight — margin", "0.15", "_similarity",
         "max(0, 1 − 5·|margin_t − margin_p|)",
         "margin proxies business model quality; 20pt gap → 0"),
        ("Similarity weight — customer", "0.15", "_similarity",
         "same B2B/B2C/mixed = 1.0 else 0", "different buyers = different economics"),
        ("Similarity weight — export", "0.10", "_similarity",
         "same exporter flag = 1.0 else 0.3",
         "the only trade signal in the data; soft because it's binary"),
        ("Match-weight taper", "floor 0.15 · full match ≥ 0.85",
         "_match_weight", "clamp((score − 0.40)/0.45, 0.15, 1.0)",
         "borderline comps inform but cannot dominate"),
        ("Peers used (top-N)", "15", "compute_valuation(top_n)",
         "highest-scoring survivors enter the multiple math",
         "industry practice: 10–20 comps"),
        ("Min multiples per method", "3", "compute_valuation",
         "method skipped (METHOD_SKIPPED_THIN) below this",
         "a 2-point distribution has no meaningful percentile"),
        ("Positioning window", "±20 pts (±30 when effective peers < 10)",
         "compute_valuation", "band around the positioned percentile",
         "thin/borderline peer sets earn wider honesty bands"),
        ("Positioning clamps", "mid_q ∈ [0.15, 0.85]; band ∈ [0.05, 0.95]",
         "compute_valuation", "never price at the extreme tails",
         "protects against outlier-driven positions"),
        ("Outlier fence", "Tukey 1.5 × IQR (skipped when < 4 values)",
         "_tukey_trim_pairs", "values outside [Q1 − 1.5·IQR, Q3 + 1.5·IQR] dropped",
         "standard robust-statistics fence"),
        ("Market-basis threshold", "≥ 3 listed comps per method",
         "compute_valuation", "else book pool + sector calibration",
         "trading multiples need real market EVs"),
        ("DLOM (private targets only)", "30% / 25% / 20%",
         "compute_valuation", "revenue < ₹100 Cr / < ₹500 Cr / ≥ ₹500 Cr; listed = 0%",
         "restricted-stock & pre-IPO studies; size-scaled marketability discount"),
        ("Scale-mismatch penalty (#15)", "−7.5%/decade below band (cap 15%) · "
         "+5%/decade above (cap 10%)", "compute_valuation",
         "band = [0.8 × smallest peer, 1.25 × largest peer]",
         "size-premium evidence: multiples are not size-transferable"),
        ("Control premium (transactions view)", "20% / 25% / 30%",
         "compute_valuation", "acquisition range = equity × (1 + premium)",
         "empirical control-premium studies; derived view, disclosed"),
        ("Confidence weights", "0.20 profile · 0.20 coverage · 0.10 EBITDA>0 · "
         "0.10 methods · 0.25 triangulation · 0.15 tightness",
         "run.py compute_confidence",
         "triangulation = 1 − spread/80% · tightness = 1 − CV/45%",
         "output coherence dominates: disagreeing methods = less trust"),
        ("Confidence labels", "HIGH ≥ 0.75 · MEDIUM ≥ 0.50 · else LOW",
         "compute_confidence", "label on every report", "three-tier communication"),
        ("Data-quality weights", "revenue 0.50 · ebitda 0.15 · cap employed 0.15 · "
         "margin 0.07 · ebit 0.05 · naics 0.05 · prior rev 0.03 · cin 0.01",
         "validate_company", "grade A ≥ 0.85 · B ≥ 0.70 · C ≥ 0.50 · else D",
         "revenue is critical; identity fields informational"),
        ("Valuable gate", "revenue OK AND (capital employed OK OR ebitda OK)",
         "validate_company", "else status insufficient_data, no valuation",
         "never fabricate from nothing"),
        ("ETL reconciliation tolerance", "max(2% × |PBDT|, ₹0.5 Cr)",
         "etl.py", "EBITDA + OI − Interest ≈ PBDT", "allows rounding noise, catches "
         "real mis-mapping"),
        ("Headline order", "EV/EBITDA → EV/Revenue → EV/EBIT",
         "compute_valuation", "first computable wins (FALLBACK_HEADLINE if not first)",
         "EBITDA multiples are the least distorted for industrials"),
        ("Sector calibration factor clamp", "[0.25, 25]",
         "compute_valuation", "anchor ÷ book median, clamped",
         "guards against degenerate book distributions"),
        ("Anchor size factor", " · ".join(f"rev < ₹{int(t)} Cr → ×{f}"
                                          for t, f in _SIZE_FACTORS) + " · else ×1.00",
         "core/calibration.py", "haircuts the sector anchor for small targets",
         "published aggregates skew large-cap; small-caps trade at a discount"),
        ("EV/EBIT anchor uplift", f"EV/EBITDA anchor × {_EBIT_UPLIFT}",
         "core/calibration.py", "D&A ≈ 20% of EBITDA for Indian industrials",
         "keeps the three methods internally consistent"),
    ]
    for grp, (ee, es) in sorted(SECTOR_ANCHORS.items()):
        par.append((f"Sector anchor — {grp}", f"EV/EBITDA {ee}x · EV/Sales {es}x",
                    "core/calibration.py SECTOR_ANCHORS",
                    "re-levels book multiples when < 3 listed comps exist",
                    "India sector aggregates (Damodaran, Jan-2025) — editable "
                    "calibration defaults; replace with a live market feed for "
                    "exact levels"))
    sheet(wb, "4 · Model parameters",
          ["Parameter", "Value", "Code location", "Mechanics", "Rationale"],
          [34, 40, 26, 52, 52], par,
          "Every constant in the engine. Change any of these in ONE place (the code "
          "location column) and the whole pipeline — CLI, API, UI, reports — picks it "
          "up. The sector anchors are calibration defaults standing in for the missing "
          "market-price feed; they are the honest, disclosed bridge from book to "
          "market levels.")

    # ---- 5 · Calculation flow ----------------------------------------------
    flow = [
        ("1 · Resolve", "company name → best match (confidence-ranked) → identifier",
         "no match → status no_match, honest stop (audit: NO_MATCH)"),
        ("2 · Normalize", "D&B envelope → Company dataclass, Thousand → Crore, "
         "EBIT/margins/growth derived", "every monetary field null-safe"),
        ("3 · Classify", "business description → operating model / value chain / "
         "customer type (manufacturing-VERB collision rule)",
         "the seam where an LLM classifier could drop in — currently rule-based"),
        ("4 · Data-quality gate", "8 weighted field checks → grade A–D + valuable flag",
         "not valuable → status insufficient_data, no number invented"),
        ("5 · Universe", "all 14,036 valuation-grade companies normalized + profiled "
         "(cached per data source; ~4s first build)", "target excluded from its own pool"),
        ("6 · Knock-outs (B)", "operating model + value chain + major industry must "
         "equal the target's", "each rejection recorded with a reason (PEER_REJECTED)"),
        ("7 · Similarity (C)", "5-dimension weighted score per survivor → ranked",
         "top 15 proceed; all survivors reported for honesty"),
        ("8 · Multiples (D)", "market pool if ≥3 listed comps, else book pool "
         "re-levelled to SECTOR ANCHORS · Tukey 1.5×IQR trim · similarity-weighted",
         "peer basis + calibration factor disclosed per method"),
        ("9 · Position", "target's EBITDA-margin percentile among peers (clamped "
         "15–85) → central multiple; ±20/±30pt band", "never sees the target's own "
         "market value — the cross-check stays independent"),
        ("10 · Adjust", "scale-mismatch penalty if target is outside the peer revenue "
         "band (filter #15)", "explicit ±% on all multiples, audited"),
        ("11 · Bridge", "EV = multiple × driver → − net debt → − DLOM (private only) "
         "→ EQUITY low/mid/high per method", "headline = first computable of "
         "EV/EBITDA → EV/Revenue → EV/EBIT"),
        ("12 · Transactions view", "indicative acquisition range = equity × "
         "(1 + 20–30% control premium)", "derived, labelled, replaceable by a deal DB"),
        ("13 · Confidence", "six weighted signals incl. method agreement + comp "
         "dispersion → score + label + breakdown", "spans ~0.31–0.95 across the "
         "universe — it discriminates"),
        ("14 · Evidence", "typed audit trail (every decision) + per-field lineage "
         "(file + row) + caveats into result.json / API / report",
         "the trust mechanism of a touchless system"),
    ]
    sheet(wb, "5 · Calculation flow",
          ["Step", "What happens", "Guarantees / notes"],
          [22, 72, 62], flow,
          "The end-to-end chain from input to answer. Every step is audited; every "
          "fallback is a recorded DECISION, never a silent default.")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    return OUT


if __name__ == "__main__":
    build()
